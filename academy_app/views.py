import os
import io
import csv
import zipfile
import datetime
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q, Count
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from academy_app.models import Academy, Trainer, Course, Batch, Student, PaymentEntry, Notification, SystemSetting, AuditLog

def parse_dmy_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.datetime.strptime(date_str.strip(), "%d-%m-%Y").date()
    except ValueError:
        try:
            return datetime.datetime.strptime(date_str.strip(), "%Y-%m-%d").date()
        except ValueError:
            return timezone.localdate()


# Helper context processor logic to pass header elements in each view
def get_header_context(request):
    selected_academy_id = request.session.get('selected_academy_id')
    active_academy_id = None
    if selected_academy_id:
        try:
            active_academy_id = int(selected_academy_id)
        except ValueError:
            pass

    header_academies = Academy.objects.filter(status='Active')
    header_notifications = Notification.objects.all().order_by('-created_at')[:10]
    unread_notifications_count = Notification.objects.filter(is_read=False).count()
    
    return {
        'header_academies': header_academies,
        'active_academy_id': active_academy_id,
        'header_notifications': header_notifications,
        'unread_notifications_count': unread_notifications_count,
    }

# Login View
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    error_msg = None
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            auth_login(request, user)
            
            # Audit log
            AuditLog.objects.create(user=user, action="Logged In", model_name="User", object_repr=user.username)
            
            return redirect('dashboard')
        else:
            error_msg = "Invalid username or password."
            
    return render(request, 'login.html', {'error_msg': error_msg})

# Logout View
def logout_view(request):
    if request.user.is_authenticated:
        AuditLog.objects.create(user=request.user, action="Logged Out", model_name="User", object_repr=request.user.username)
    auth_logout(request)
    return redirect('login')

# Dashboard View
@login_required
def dashboard_view(request):
    # Set selected academy from GET param or session
    acad_param = request.GET.get('selected_academy_id')
    if acad_param is not None:
        if acad_param == '':
            if 'selected_academy_id' in request.session:
                del request.session['selected_academy_id']
        else:
            request.session['selected_academy_id'] = acad_param
            
    context = get_header_context(request)
    active_acad_id = context['active_academy_id']

    # Filter base querysets
    academies = Academy.objects.all()
    students = Student.objects.all()
    payments = PaymentEntry.objects.all()
    
    if active_acad_id:
        students = students.filter(academy_id=active_acad_id)
        payments = payments.filter(academy_id=active_acad_id)
        academies = academies.filter(id=active_acad_id)

    today = timezone.localdate()
    start_of_month = today.replace(day=1)
    
    # Reminders & counts
    overdue_count = 0
    upcoming_count = 0
    overdue_fees = Decimal('0.00')

    # Date range for upcoming (next 7 days)
    next_7_days = today + datetime.timedelta(days=7)
    
    # Calculate dues dynamically per student
    for s in students.filter(status='Active'):
        # Last payment balance + registration if first time
        last_payment = s.payments.order_by('id').last()
        out_bal = last_payment.balance if last_payment else s.registration_fee
        
        if s.last_due_date < today:
            overdue_count += 1
            # Add outstanding + monthly fee if overdue
            overdue_fees += out_bal + s.monthly_fee
        elif s.last_due_date <= next_7_days:
            upcoming_count += 1
            
    # Quick Stats
    today_collection = payments.filter(payment_date=today).aggregate(sum=Sum('amount_paid'))['sum'] or Decimal('0.00')
    monthly_collection = payments.filter(payment_date__gte=start_of_month).aggregate(sum=Sum('amount_paid'))['sum'] or Decimal('0.00')
    
    # Total monthly rate of all active students
    expected_collection = students.filter(status='Active').aggregate(sum=Sum('monthly_fee'))['sum'] or Decimal('0.00')
    pending_collection = max(Decimal('0.00'), expected_collection - monthly_collection)
    
    stats = {
        'total_academies': Academy.objects.count(),
        'total_students': students.count(),
        'active_students': students.filter(status='Active').count(),
        'today_collection': today_collection,
        'monthly_collection': monthly_collection,
        'pending_collection': pending_collection,
        'overdue_fees': overdue_fees,
    }

    # Chart 1: Monthly collections & dues comparison for past 6 months
    chart_months = []
    chart_collected = []
    chart_outstanding = []
    
    for i in range(5, -1, -1):
        m_date = today - datetime.timedelta(days=i*30)
        m_start = m_date.replace(day=1)
        m_label = m_date.strftime("%B")
        chart_months.append(m_label)
        
        # Monthly collected sum
        col_val = payments.filter(payment_date__year=m_date.year, payment_date__month=m_date.month).aggregate(sum=Sum('amount_paid'))['sum'] or Decimal('0.00')
        chart_collected.append(float(col_val))
        
        # Monthly outstanding sum
        out_val = payments.filter(payment_date__year=m_date.year, payment_date__month=m_date.month).aggregate(sum=Sum('balance'))['sum'] or Decimal('0.00')
        chart_outstanding.append(float(out_val))
        
    # Chart 2: Academy Wise collections
    academy_names = []
    academy_collections = []
    
    for acad in Academy.objects.filter(status='Active'):
        c_val = PaymentEntry.objects.filter(academy=acad).aggregate(sum=Sum('amount_paid'))['sum'] or Decimal('0.00')
        academy_names.append(acad.name)
        academy_collections.append(float(c_val))

    # General Settings Check
    fy_setting = SystemSetting.objects.filter(key='financial_year').first()
    current_fy = fy_setting.value if fy_setting else "2026-2027"

    # Recent payments
    recent_payments = payments.order_by('-id')[:5]
    
    # Active notifications
    dashboard_notifications = Notification.objects.filter(is_read=False).order_by('-created_at')[:5]

    # Automatic Reminder alerts generation on dashboard load
    # Look for students overdue
    for s in students.filter(status='Active', last_due_date__lt=today)[:5]:
        title = f"Fee Overdue: {s.name}"
        msg = f"Student {s.name} ({s.student_id}) has crossed renewal date {s.last_due_date.strftime('%d %M %Y')}."
        if not Notification.objects.filter(title=title).exists():
            Notification.objects.create(title=title, message=msg, notification_type='Alert')

    context.update({
        'stats': stats,
        'overdue_count': overdue_count,
        'upcoming_count': upcoming_count,
        'current_fy': current_fy,
        'recent_payments': recent_payments,
        'dashboard_notifications': dashboard_notifications,
        'chart_months': chart_months,
        'chart_collected': chart_collected,
        'chart_outstanding': chart_outstanding,
        'academy_names': academy_names,
        'academy_collections': academy_collections,
    })
    
    return render(request, 'dashboard.html', context)

# Academies View
@login_required
def academies_view(request):
    context = get_header_context(request)
    
    if request.method == 'POST':
        # Create or Update
        acad_id = request.POST.get('academy_id')
        name = request.POST.get('name')
        code = request.POST.get('code')
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        gst = request.POST.get('gst_number')
        status = request.POST.get('status', 'Active')
        logo = request.FILES.get('logo')

        if acad_id:
            academy = get_object_or_404(Academy, id=acad_id)
            academy.name = name
            academy.code = code
            academy.address = address
            academy.phone = phone
            academy.email = email
            academy.gst_number = gst
            academy.status = status
            if logo:
                academy.logo = logo
            academy.save()
            AuditLog.objects.create(user=request.user, action="Updated Academy", model_name="Academy", object_id=academy.id, object_repr=academy.name)
        else:
            academy = Academy.objects.create(
                name=name, code=code, address=address, phone=phone, email=email, gst_number=gst, status=status, logo=logo
            )
            AuditLog.objects.create(user=request.user, action="Created Academy", model_name="Academy", object_id=academy.id, object_repr=academy.name)
            
        return redirect('academies')

    academies = Academy.objects.all().order_by('-id')
    context.update({
        'academies': academies,
    })
    return render(request, 'academies.html', context)

# Students View
@login_required
def students_view(request):
    context = get_header_context(request)
    active_acad_id = context['active_academy_id']

    if request.method == 'POST':
        student_db_id = request.POST.get('student_db_id')
        admission_no = request.POST.get('admission_no')
        name = request.POST.get('name')
        gender = request.POST.get('gender')
        dob = request.POST.get('dob')
        parent_name = request.POST.get('parent_name')
        mobile = request.POST.get('mobile')
        alt_mobile = request.POST.get('alt_mobile')
        email = request.POST.get('email')
        address = request.POST.get('address')
        
        academy_id = request.POST.get('academy')
        course_id = request.POST.get('course')
        batch_id = request.POST.get('batch')
        
        monthly_fee = request.POST.get('monthly_fee')
        registration_fee = request.POST.get('registration_fee', 0.00)
        discount = request.POST.get('discount', 0.00)
        joining_date = request.POST.get('joining_date')
        status = request.POST.get('status', 'Active')
        photo = request.FILES.get('photo')

        # Resolve objects
        academy = get_object_or_404(Academy, id=academy_id)
        course = get_object_or_404(Course, id=course_id)
        batch = get_object_or_404(Batch, id=batch_id)
        trainer = batch.trainer

        parsed_dob = parse_dmy_date(dob)
        parsed_joining_date = parse_dmy_date(joining_date)

        if student_db_id:
            student = get_object_or_404(Student, id=student_db_id)
            student.admission_no = admission_no
            student.name = name
            student.gender = gender
            student.dob = parsed_dob
            student.parent_name = parent_name
            student.mobile = mobile
            student.alt_mobile = alt_mobile
            student.email = email
            student.address = address
            student.academy = academy
            student.course = course
            student.batch = batch
            student.trainer = trainer
            student.monthly_fee = monthly_fee
            student.registration_fee = registration_fee
            student.discount = discount
            student.joining_date = parsed_joining_date
            student.status = status
            if photo:
                student.photo = photo
            student.save()
            AuditLog.objects.create(user=request.user, action="Updated Student Profile", model_name="Student", object_id=student.id, object_repr=student.name)
        else:
            student = Student.objects.create(
                admission_no=admission_no, name=name, gender=gender, dob=parsed_dob, parent_name=parent_name,
                mobile=mobile, alt_mobile=alt_mobile, email=email, address=address, academy=academy,
                course=course, batch=batch, trainer=trainer, monthly_fee=monthly_fee,
                registration_fee=registration_fee, discount=discount, joining_date=parsed_joining_date, status=status, photo=photo
            )
            Notification.objects.create(
                title="New Admission",
                message=f"{student.name} was admitted to {academy.name} under {course.name}.",
                notification_type='Info'
            )
            AuditLog.objects.create(user=request.user, action="Created Student Record", model_name="Student", object_id=student.id, object_repr=student.name)

        return redirect('students')

    # Filtering
    students_qs = Student.objects.all().order_by('-id')
    if active_acad_id:
        students_qs = students_qs.filter(academy_id=active_acad_id)

    search_q = request.GET.get('search', '')
    course_id_q = request.GET.get('course_id', '')
    trainer_id_q = request.GET.get('trainer_id', '')

    if search_q:
        students_qs = students_qs.filter(
            Q(name__icontains=search_q) | Q(student_id__icontains=search_q) | 
            Q(admission_no__icontains=search_q) | Q(parent_name__icontains=search_q)
        )
    if course_id_q:
        students_qs = students_qs.filter(course_id=course_id_q)
    if trainer_id_q:
        students_qs = students_qs.filter(trainer_id=trainer_id_q)

    academies_list = Academy.objects.filter(status='Active')
    all_courses = Course.objects.all()
    all_trainers = Trainer.objects.all()
    
    if active_acad_id:
        all_courses = all_courses.filter(academy_id=active_acad_id)
        all_trainers = all_trainers.filter(academy_id=active_acad_id)

    context.update({
        'students': students_qs,
        'academies_list': academies_list,
        'all_courses': all_courses,
        'all_trainers': all_trainers,
        'today': timezone.localdate(),
        'next_week': timezone.localdate() + datetime.timedelta(days=7),
        'filters': {
            'search': search_q,
            'course_id': course_id_q,
            'trainer_id': trainer_id_q,
        }
    })
    return render(request, 'students.html', context)

# Single screen Payment Desk View
@login_required
def payment_desk_view(request):
    context = get_header_context(request)
    academies_list = Academy.objects.filter(status='Active')
    
    context.update({
        'academies_list': academies_list,
    })
    return render(request, 'payment_desk.html', context)

# API: Auto-Complete Student search (Instant 3 letters)
@login_required
def api_student_search(request):
    query = request.GET.get('q', '')
    academy_id = request.GET.get('academy_id', '')

    if len(query) < 3:
        return JsonResponse({'results': []})

    qs = Student.objects.filter(status='Active')
    if academy_id:
        qs = qs.filter(academy_id=academy_id)

    qs = qs.filter(
        Q(name__icontains=query) | Q(student_id__icontains=query) | Q(admission_no__icontains=query)
    )[:10]

    results = []
    for s in qs:
        results.append({
            'id': s.id,
            'name': s.name,
            'student_id': s.student_id,
            'admission_no': s.admission_no,
            'course': s.course.name
        })

    return JsonResponse({'results': results})

# API: Get student full profile details for payment workspace loading
@login_required
def api_student_details(request, pk):
    student = get_object_or_404(Student, id=pk)
    
    # Calculate outstanding balance
    last_payment = student.payments.order_by('id').last()
    outstanding_balance = last_payment.balance if last_payment else student.registration_fee

    today = timezone.localdate()
    due_status = "Paid"
    if student.last_due_date < today:
        due_status = "Overdue"
    elif student.last_due_date <= today + datetime.timedelta(days=7):
        due_status = "Due Soon"

    # Photo URL
    photo_url = student.photo.url if student.photo else ""

    student_data = {
        'id': student.id,
        'name': student.name,
        'student_id': student.student_id,
        'admission_no': student.admission_no,
        'dob': student.dob.strftime('%d-%m-%Y'),
        'parent_name': student.parent_name,
        'mobile': student.mobile,
        'alt_mobile': student.alt_mobile,
        'email': student.email,
        'address': student.address,
        'academy_id': student.academy.id,
        'academy': student.academy.name,
        'course_id': student.course.id,
        'course': student.course.name,
        'batch_id': student.batch.id,
        'batch': student.batch.name,
        'trainer': student.trainer.name,
        'monthly_fee': str(student.monthly_fee),
        'registration_fee': str(student.registration_fee),
        'discount': str(student.discount),
        'joining_date': student.joining_date.strftime('%d-%m-%Y'),
        'status': student.status,
        'last_due_date': student.last_due_date.strftime('%d-%m-%Y'),
        'last_due_date_formatted': student.last_due_date.strftime('%d-%m-%Y'),
        'outstanding_balance': str(outstanding_balance),
        'due_status': due_status,
        'photo_url': photo_url,
    }

    # Fetch past payments list
    payments_qs = student.payments.all().order_by('-id')
    payments_data = []
    for p in payments_qs:
        payments_data.append({
            'id': p.id,
            'receipt_no': p.receipt_no,
            'payment_date': p.payment_date.strftime('%d-%m-%Y'),
            'payment_period': p.payment_period,
            'amount_paid': str(p.amount_paid),
            'balance': str(p.balance),
            'status': p.status,
            'remarks': p.remarks,
        })

    return JsonResponse({
        'student': student_data,
        'payments': payments_data
    })

# API: Dynamic Settings Metadata lists for drop-down chaining (Academy -> Course -> Batch)
@login_required
def api_settings_metadata(request):
    courses = {}
    batches = {}

    for acad in Academy.objects.filter(status='Active'):
        courses[acad.id] = []
        for c in acad.courses.all():
            courses[acad.id].append({'id': c.id, 'name': c.name})

    for course in Course.objects.all():
        batches[course.id] = []
        for b in course.batches.all():
            batches[course.id].append({
                'id': b.id,
                'name': b.name,
                'trainer': b.trainer.name
            })

    return JsonResponse({
        'courses': courses,
        'batches': batches
    })

# Payment Create View (AJAX Handler)
@login_required
def payment_create_view(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        academy_id = request.POST.get('academy_id')
        payment_date = request.POST.get('payment_date')
        payment_period = request.POST.get('payment_period')
        amount_due = Decimal(request.POST.get('amount_due', 0))
        amount_paid = Decimal(request.POST.get('amount_paid', 0))
        discount = Decimal(request.POST.get('discount', 0))
        balance = Decimal(request.POST.get('balance', 0))
        payment_mode = request.POST.get('payment_mode', 'Cash')
        transaction_id = request.POST.get('transaction_id')
        next_due_date = request.POST.get('next_due_date')
        remarks = request.POST.get('remarks')

        student = get_object_or_404(Student, id=student_id)
        academy = get_object_or_404(Academy, id=academy_id)

        # Status categorization (Paid or Partial Payment)
        status = 'Paid'
        if balance > 0:
            status = 'Partial Payment'

        parsed_payment_date = parse_dmy_date(payment_date)
        parsed_next_due_date = parse_dmy_date(next_due_date)

        payment = PaymentEntry.objects.create(
            student=student,
            academy=academy,
            payment_date=parsed_payment_date,
            payment_period=payment_period,
            amount_due=amount_due,
            amount_paid=amount_paid,
            discount=discount,
            balance=balance,
            payment_mode=payment_mode,
            transaction_id=transaction_id,
            next_due_date=parsed_next_due_date,
            remarks=remarks,
            status=status,
            created_by=request.user
        )

        # Update student record validity
        student.last_due_date = parsed_next_due_date
        student.save()

        # Audit logs & Dashboard alerts
        AuditLog.objects.create(
            user=request.user,
            action="Created Payment Log",
            model_name="PaymentEntry",
            object_id=payment.id,
            object_repr=payment.receipt_no
        )

        Notification.objects.create(
            title="Payment Received",
            message=f"Received ₹{amount_paid} from {student.name} ({payment.receipt_no}). Next Due: {next_due_date}.",
            notification_type='Success'
        )

        return JsonResponse({
            'success': True,
            'payment_id': payment.id,
            'receipt_no': payment.receipt_no
        })

    return JsonResponse({'success': False, 'message': 'Invalid HTTP Request method.'})

# Printable Receipt layout View
@login_required
def receipt_pdf_view(request, pk):
    payment = get_object_or_404(PaymentEntry, id=pk)
    return render(request, 'receipt_print.html', {'payment': payment})

# Collection Reports View
@login_required
def reports_view(request):
    context = get_header_context(request)
    active_acad_id = context['active_academy_id']

    # Filters parsing
    rep_academy_id = request.GET.get('report_academy_id') or active_acad_id
    rep_course_id = request.GET.get('report_course_id')
    rep_trainer_id = request.GET.get('report_trainer_id')
    rep_payment_mode = request.GET.get('report_payment_mode')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    rep_status = request.GET.get('report_status')

    payments_qs = PaymentEntry.objects.all().order_by('-id')

    if rep_academy_id:
        payments_qs = payments_qs.filter(academy_id=rep_academy_id)
    if rep_course_id:
        payments_qs = payments_qs.filter(student__course_id=rep_course_id)
    if rep_trainer_id:
        payments_qs = payments_qs.filter(student__trainer_id=rep_trainer_id)
    if rep_payment_mode:
        payments_qs = payments_qs.filter(payment_mode=rep_payment_mode)
    if date_start:
        payments_qs = payments_qs.filter(payment_date__gte=parse_dmy_date(date_start))
    if date_end:
        payments_qs = payments_qs.filter(payment_date__lte=parse_dmy_date(date_end))
    if rep_status:
        payments_qs = payments_qs.filter(status=rep_status)

    # Summaries calculator
    collected = payments_qs.aggregate(sum=Sum('amount_paid'))['sum'] or Decimal('0.00')
    outstanding = payments_qs.aggregate(sum=Sum('balance'))['sum'] or Decimal('0.00')
    concessions = payments_qs.aggregate(sum=Sum('discount'))['sum'] or Decimal('0.00')
    expected = collected + outstanding + concessions
    
    rate = 0.0
    if expected > 0:
        rate = float(collected / expected) * 100

    summary = {
        'expected': expected,
        'collected': collected,
        'outstanding': outstanding,
        'rate': rate,
    }

    # Meta filters data loading
    academies_list = Academy.objects.filter(status='Active')
    all_courses = Course.objects.all()
    all_trainers = Trainer.objects.all()
    if active_acad_id:
        all_courses = all_courses.filter(academy_id=active_acad_id)
        all_trainers = all_trainers.filter(academy_id=active_acad_id)

    context.update({
        'collections': payments_qs,
        'summary': summary,
        'academies_list': academies_list,
        'all_courses': all_courses,
        'all_trainers': all_trainers,
        'filters': {
            'report_academy_id': rep_academy_id,
            'report_course_id': rep_course_id,
            'report_trainer_id': rep_trainer_id,
            'report_payment_mode': rep_payment_mode,
            'date_start': date_start or '',
            'date_end': date_end or '',
            'report_status': rep_status,
        }
    })
    return render(request, 'reports.html', context)

# Reports CSV Export
@login_required
def export_csv_view(request):
    rep_academy_id = request.GET.get('report_academy_id')
    rep_course_id = request.GET.get('report_course_id')
    rep_trainer_id = request.GET.get('report_trainer_id')
    rep_payment_mode = request.GET.get('report_payment_mode')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    rep_status = request.GET.get('report_status')

    payments_qs = PaymentEntry.objects.all().order_by('-id')

    if rep_academy_id:
        payments_qs = payments_qs.filter(academy_id=rep_academy_id)
    if rep_course_id:
        payments_qs = payments_qs.filter(student__course_id=rep_course_id)
    if rep_trainer_id:
        payments_qs = payments_qs.filter(student__trainer_id=rep_trainer_id)
    if rep_payment_mode:
        payments_qs = payments_qs.filter(payment_mode=rep_payment_mode)
    if date_start:
        payments_qs = payments_qs.filter(payment_date__gte=parse_dmy_date(date_start))
    if date_end:
        payments_qs = payments_qs.filter(payment_date__lte=parse_dmy_date(date_end))
    if rep_status:
        payments_qs = payments_qs.filter(status=rep_status)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="fee_collection_report.csv"'

    writer = csv.writer(response)
    writer.writerow(['Receipt No', 'Payment Date', 'Student Name', 'Academy', 'Course', 'Period', 'Amount Due', 'Amount Paid', 'Discount', 'Balance', 'Payment Mode', 'Transaction ID'])

    for p in payments_qs:
        writer.writerow([
            p.receipt_no, p.payment_date, p.student.name, p.academy.name,
            p.student.course.name, p.payment_period, p.amount_due, p.amount_paid,
            p.discount, p.balance, p.payment_mode, p.transaction_id
        ])

    return response

# Reports Excel Export using openpyxl
@login_required
def export_excel_view(request):
    rep_academy_id = request.GET.get('report_academy_id')
    rep_course_id = request.GET.get('report_course_id')
    rep_trainer_id = request.GET.get('report_trainer_id')
    rep_payment_mode = request.GET.get('report_payment_mode')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    rep_status = request.GET.get('report_status')

    payments_qs = PaymentEntry.objects.all().order_by('-id')

    if rep_academy_id:
        payments_qs = payments_qs.filter(academy_id=rep_academy_id)
    if rep_course_id:
        payments_qs = payments_qs.filter(student__course_id=rep_course_id)
    if rep_trainer_id:
        payments_qs = payments_qs.filter(student__trainer_id=rep_trainer_id)
    if rep_payment_mode:
        payments_qs = payments_qs.filter(payment_mode=rep_payment_mode)
    if date_start:
        payments_qs = payments_qs.filter(payment_date__gte=parse_dmy_date(date_start))
    if date_end:
        payments_qs = payments_qs.filter(payment_date__lte=parse_dmy_date(date_end))
    if rep_status:
        payments_qs = payments_qs.filter(status=rep_status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Collections Audit"

    # Styling
    font_header = Font(name="Outfit", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    headers = ['Receipt No', 'Payment Date', 'Student Name', 'Academy', 'Course', 'Period', 'Amount Due', 'Amount Paid', 'Discount', 'Balance', 'Payment Mode', 'Transaction ID']
    ws.append(headers)

    # Style header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for p in payments_qs:
        ws.append([
            p.receipt_no,
            p.payment_date.strftime("%Y-%m-%d"),
            p.student.name,
            p.academy.name,
            p.student.course.name,
            p.payment_period,
            float(p.amount_due),
            float(p.amount_paid),
            float(p.discount),
            float(p.balance),
            p.payment_mode,
            p.transaction_id or ""
        ])

    # Auto fit column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fee_collection_audit.xlsx"'
    wb.save(response)
    
    return response

# Configuration / Settings Page View
@login_required
def settings_view(request):
    context = get_header_context(request)
    
    # Load all settings
    keys = ['receipt_prefix', 'financial_year', 'sms_enabled', 'email_enabled', 'whatsapp_enabled']
    sys_settings = {}
    for key in keys:
        setting = SystemSetting.objects.filter(key=key).first()
        sys_settings[key] = setting.value if setting else "False"

    academies_list = Academy.objects.filter(status='Active')
    courses_list = Course.objects.all()
    trainers_list = Trainer.objects.all()
    batches_list = Batch.objects.all()

    context.update({
        'sys_settings': sys_settings,
        'academies_list': academies_list,
        'courses_list': courses_list,
        'trainers_list': trainers_list,
        'batches_list': batches_list,
    })
    return render(request, 'settings.html', context)

# General Settings update view
@login_required
def update_settings_view(request):
    if request.method == 'POST':
        keys = ['receipt_prefix', 'financial_year', 'sms_enabled', 'email_enabled', 'whatsapp_enabled']
        for key in keys:
            val = request.POST.get(key, 'False')
            # Checkbox values are handled
            if key in ['sms_enabled', 'email_enabled', 'whatsapp_enabled'] and val != 'True':
                val = 'False'
            
            setting, created = SystemSetting.objects.get_or_create(key=key)
            setting.value = val
            setting.save()

        AuditLog.objects.create(user=request.user, action="Updated General System Settings", model_name="SystemSetting", object_repr="System Settings")
        return redirect('settings')
    return HttpResponse("Invalid request.")

# Create settings objects directly from config panel
@login_required
def create_course_view(request):
    if request.method == 'POST':
        c_name = request.POST.get('course_name')
        acad_id = request.POST.get('course_academy_id')
        academy = get_object_or_404(Academy, id=acad_id)
        Course.objects.create(name=c_name, academy=academy)
        return redirect('settings')

@login_required
def create_batch_view(request):
    if request.method == 'POST':
        b_name = request.POST.get('batch_name')
        c_id = request.POST.get('batch_course_id')
        t_id = request.POST.get('batch_trainer_id')
        course = get_object_or_404(Course, id=c_id)
        trainer = get_object_or_404(Trainer, id=t_id)
        Batch.objects.create(name=b_name, course=course, trainer=trainer)
        return redirect('settings')

@login_required
def create_trainer_view(request):
    if request.method == 'POST':
        t_name = request.POST.get('trainer_name')
        t_email = request.POST.get('trainer_email')
        t_phone = request.POST.get('trainer_phone')
        t_spec = request.POST.get('trainer_specialization')
        acad_id = request.POST.get('trainer_academy_id')
        academy = get_object_or_404(Academy, id=acad_id)
        Trainer.objects.create(name=t_name, email=t_email, phone=t_phone, specialization=t_spec, academy=academy)
        return redirect('settings')

# Mark all notification items as read view
@login_required
def mark_notifications_read_view(request):
    Notification.objects.filter(is_read=False).update(is_read=True)
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

# Database Backup (Manual zip download)
@login_required
def trigger_backup_view(request):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_path = os.path.join(base_dir, 'db.sqlite3')
    
    if not os.path.exists(db_path):
        return HttpResponse("Database file not found.")

    # Create memory ZIP
    byte_stream = io.BytesIO()
    with zipfile.ZipFile(byte_stream, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.write(db_path, arcname='db.sqlite3')

    byte_stream.seek(0)
    
    response = HttpResponse(byte_stream.read(), content_type='application/zip')
    stamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="feeflow_backup_{stamp}.zip"'
    
    AuditLog.objects.create(user=request.user, action="Created System Database Backup", model_name="Database", object_repr="db.sqlite3")
    return response

# Database Restore (Manual zip upload)
@login_required
def restore_backup_view(request):
    if request.method == 'POST':
        b_file = request.FILES.get('backup_file')
        if not b_file:
            return HttpResponse("No file uploaded.")

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        db_path = os.path.join(base_dir, 'db.sqlite3')

        # Check file type
        filename = b_file.name
        if filename.endswith('.zip'):
            # Extract zip in memory and overwrite
            try:
                with zipfile.ZipFile(b_file) as zip_ref:
                    # Look for db.sqlite3 inside zip
                    if 'db.sqlite3' in zip_ref.namelist():
                        data = zip_ref.read('db.sqlite3')
                        with open(db_path, 'wb') as dest:
                            dest.write(data)
                    else:
                        return HttpResponse("ZIP file does not contain db.sqlite3.")
            except Exception as e:
                return HttpResponse(f"Error extracting ZIP file: {str(e)}")
        elif filename.endswith('.db') or filename.endswith('.sqlite3'):
            # Replace database directly
            try:
                with open(db_path, 'wb+') as destination:
                    for chunk in b_file.chunks():
                        destination.write(chunk)
            except Exception as e:
                return HttpResponse(f"Error replacing database file: {str(e)}")
        else:
            return HttpResponse("Unsupported file format. Upload .zip, .db, or .sqlite3.")

        AuditLog.objects.create(user=request.user, action="Restored System Database Backup", model_name="Database", object_repr="db.sqlite3")
        
        # Log back in user or redirect
        return redirect('dashboard')
    
    return HttpResponse("Invalid request.")

# Import Students from Excel Sheet
@login_required
def import_students_view(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return HttpResponse("No file provided.")
            
        import openpyxl
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Row 1 is header
            headers = [cell.value for cell in ws[1]]
            
            # Required keys mapping
            # Index check
            row_count = 0
            for r_idx in range(2, ws.max_row + 1):
                row = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, ws.max_column + 1)]
                if not row or not any(row):
                    continue
                
                # Zip headers & values
                r_data = dict(zip(headers, row))
                
                # Fetch objects
                acad_code = r_data.get('Academy Code')
                c_name = r_data.get('Course Name')
                b_name = r_data.get('Batch Name')
                
                academy = Academy.objects.filter(code=acad_code).first()
                if not academy:
                    continue
                    
                course = Course.objects.filter(name=c_name, academy=academy).first()
                if not course:
                    course = Course.objects.create(name=c_name, academy=academy)
                    
                # Trainer fallback
                trainer = Trainer.objects.filter(academy=academy).first()
                if not trainer:
                    trainer = Trainer.objects.create(name="Default Coach", email="coach@example.com", phone="0000000000", academy=academy)
                    
                batch = Batch.objects.filter(name=b_name, course=course).first()
                if not batch:
                    batch = Batch.objects.create(name=b_name, course=course, trainer=trainer)
                
                # Convert date formats safely
                def parse_date(date_val):
                    if isinstance(date_val, datetime.date):
                        return date_val
                    elif isinstance(date_val, datetime.datetime):
                        return date_val.date()
                    elif isinstance(date_val, str):
                        try:
                            return datetime.datetime.strptime(date_val, "%Y-%m-%d").date()
                        except:
                            return timezone.localdate()
                    return timezone.localdate()

                dob = parse_date(r_data.get('DOB'))
                join_date = parse_date(r_data.get('Joining Date'))

                # Check duplicate admission_no
                adm_no = str(r_data.get('Admission No', ''))
                if Student.objects.filter(admission_no=adm_no).exists():
                    continue

                Student.objects.create(
                    admission_no=adm_no,
                    name=r_data.get('Student Name'),
                    gender=r_data.get('Gender', 'Male'),
                    dob=dob,
                    parent_name=r_data.get('Parent Name', 'Parent'),
                    mobile=str(r_data.get('Mobile', '0000000000')),
                    email=r_data.get('Email', 'student@example.com'),
                    address=r_data.get('Address', 'Address'),
                    joining_date=join_date,
                    academy=academy,
                    course=course,
                    batch=batch,
                    trainer=trainer,
                    monthly_fee=Decimal(str(r_data.get('Monthly Fee', 3000.00))),
                    registration_fee=Decimal('0.00'),
                    discount=Decimal('0.00'),
                    status='Active'
                )
                row_count += 1
                
            AuditLog.objects.create(user=request.user, action=f"Imported {row_count} Students from Excel", model_name="Student", object_repr="Excel File")
        except Exception as e:
            return HttpResponse(f"Error parsing Excel: {str(e)}")

        return redirect('students')
    return HttpResponse("Invalid request.")
